# -*- coding: utf-8 -*-
"""
Created on Fri Aug 21 20:31:39 2020

@author: VINOD MAHESH JAIN
"""

# Importing required libraries into envirnoment
import requests
import openpyxl as op
import time

# weather_data() is used to get the json object from the openweathermap.org of the provided city
def weather_data(query):
	return requests.get(query).json()

# returns the temperature of the provided city
def getTemperature(city):
    weather_report = weather_data(base+'q='+city+'&APPID='+AppID+'&units=metric')
    return weather_report['main']['temp']

# Building queries to connect to the openweathermap.ord
base = 'http://api.openweathermap.org/data/2.5/weather?'
AppID = '3704ad9fc427800b60c884a0de23fda4'

# Loading the excel sheet
weather = op.load_workbook('Weather_Info.xlsx')
# Getting access to it sheets to update
w = weather.active
w1 = weather['Weather']
w2  = weather['City Tokens']

# dic is used to maintain the key value pairs for the city name and city token
dic = {}
for i in range(2,15495):
    dic[w2.cell(row=i,column=2).value] = w2.cell(row=i,column=1).value
    
# As it should continuines update the sheet put it under while
while True:
    for i in range(2,10):
        # Verify whether to update current row using the feature 'Update(0/1)' i.e., 0 means stop and 1 means update        
        if int(w1.cell(row=i, column=4).value) == 1:
            # Get the city name from the excel
            city = dic[w1.cell(row=i,column=1).value]
            # call getTemperature() to get temperature of the city provided
            temp = getTemperature(city)
            # Check for the metric to present the data in i.e., in celsius or fahrenheit using the feature 'Unit(C/F)' i.e., C means celsius and F means fahrenheit
            if str(w1.cell(row=i, column=3).value) == "C":
                w1.cell(row=i, column=2).value = temp
            elif str(w1.cell(row=i, column=3).value) == "F":
                # Convert celsius into fahrenheit
                w1.cell(row=i, column=2).value = ((temp*1.8) + 32)
    # Save the weather info in excel
    weather.save('Weather_Info.xlsx')
    # Wait for 5 second and continue again
    time.sleep(5)
print("Exit")